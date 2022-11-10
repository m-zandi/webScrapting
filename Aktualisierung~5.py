
import requests
import json
import sys

#//
import time
import datetime
import convertdate
from convertdate import persian
from datetime import timedelta
import calendar

#///scraping library
# import bs4
# from urllib.request import urlopen as uReq
# from bs4 import BeautifulSoup as soup


import threading
#s
import configparser as cfg

#telegram files
import telegram

#file excel
import openpyxl
import xlrd
#email
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os.path
#array
import numpy
#my classes imported
from LinksAndContainer import arrayLinks
from LinksAndContainer import CurrentContainerClass
from todayTimeDate import todayTimeDate

def main():
    while True:
                try:
                    startTime = 5 * 60 * 60
                    finishTime = 17 * 60 * 60
                    wholeTime = 24 * 60 * 60  
                    print("startTime = ",startTime)
                    print("finishTime =",finishTime)          
                    print("now = ",nowPerSecond())

                    startApp = datetime.datetime.now()
                    # fCun = openpyxl.load_workbook("fAppendReport.xlsx")
                    # fCunSheet1 = fCun.get_sheet_by_name("firstContainerSheet")
                    cunInv = openpyxl.load_workbook("appendReport.xlsx")
                    cunSheet1 = cunInv.get_sheet_by_name("filesInvolved")
                    changesSheet = cunInv.get_sheet_by_name("changes")
                    
                    # botschaftbot = meys
                    # bot = telegram.Bot(botschaftbot)
                    ####first current others
                    # fCun = openpyxl.load_workbook("fAppendReport.xlsx")
                    # x=0
                    # fCunSheet1 = fCun.get_sheet_by_name("firstContainerSheet")
                    #####
                    objMsg =  arrayLinks().arrayMsgFunc()
                    objFA = arrayLinks().arrayAddMsgFunc()                    
                    objtTDM = todayTimeDate().tTDM()
                    #######
                    
                    ###############first file save first pdfs
                    
                    # b=0
                    # while True:
                    if cunSheet1["A1"].value  is None:
                            fCArray= arrayLinks().arrayLinksFunc()
                            for x in range(42):                            
                                        cunSheet1["A{}".format(x+1)].value = fCArray[x]
                                        cunSheet1["B{}".format(x+1)].value = " : fCArray[{}]".format(x)
                                        cunSheet1["E{}".format(x+1)].value = objtTDM
                                        cunSheet1["F{}".format(x+1)].value = objMsg[x]
                                        cunSheet1["G{}".format(x+1)].value = objFA[x]
                                        cunInv.save("appendReport.xlsx")
                        
                                
                    #########succeed save
                    #put database on array one demontial but DO NOT FORGET SET SHEETS COPY PASTE AND IF SECTION
                    book = xlrd.open_workbook('appendReport.xlsx')
                    # sheet = book.sheet_by_name('saved')
                    sheet = book.sheet_by_name('filesInvolved')
                    data = [[sheet.cell_value(r, c) for c in range(1)] for r in range(sheet.nrows)]
                    db = numpy.array(data).flatten()
                    dbFCArray = list(db)
                    # dbSaved = dbFCArray
                    # print("dbFCArray[] = ", dbFCArray)
                    # Profit !
                    # print("dbFCArray = ",dbFCArray)




                    fA = arrayLinks.fA 
                    channel = "@AWDBotschaft"
                    chAdd = "{} \n @DeutschOhal".format(channel)
                    tel =telegram_bot("config.cfg")
                    # myId = 130405462
                    # channel = myId
                    
                    try:
                        finishAppfirst =  datetime.datetime.now()
                        while True:
                                if startTime < nowPerSecond()< finishTime:
                                        #start duration time of current file and after first duration
                                        start2App = datetime.datetime.now()
                                        ############termporary 
                                        objDT = todayTimeDate().tTDM()
                                        cCArray = arrayLinks().arrayLinksFunc()
                                        # print("dbFCArray[] = ",dbFCArray)
                                        # print("cCArray  [] = ",cCArray )
                                        #finish duration time of app
                                        finishApp = datetime.datetime.now()
                                        #first run app duration time
                                        firstDurationRunAppTime = finishAppfirst - startApp
                                        #current run app duration time
                                        currentDurationRunAppTime = finishApp - start2App
                                        timeSleep = 0
                                        # fInvSheet1["A1"].value = "comparison"
                                        for x in range(42): 
                                                    #COMPARISON WITH FIRST CONTAINER
                                                    # if  fCArray[x] == cCArray[x]:
                                                    #COMPARISON BY DATABASE

                                                    
                                                    if  dbFCArray[x] == cCArray[x]:
                                                        val = "fCArray[{}] = {}\ncCArray[{}] = {}\n{}\nsame! \nFirst Duration Run App Time = {}\ndSecond Duration Run App Time = {}\n come after {} second".format(x,dbFCArray[x],x,cCArray[x],objDT,firstDurationRunAppTime,currentDurationRunAppTime,timeSleep)
                                                        print(val)
                                                        cunSheet1["C{}".format(x+1)].value = cCArray[x]
                                                        cunSheet1["D{}".format(x+1)].value = ": cCArray[{}]".format(x)
                                                        cunSheet1["E{}".format(x+1)].value = objDT
                                                        cunSheet1["F{}".format(x+1)].value = objMsg[x]
                                                        cunSheet1["G{}".format(x+1)].value = objFA[x]
                                                        cunSheet1["H{}".format(x+1)].value = "firstDurationRunAppTime = {},currentDurationRunAppTime = {}".format(firstDurationRunAppTime,currentDurationRunAppTime)
                                                        
                                                        cunInv.save("appendReport.xlsx")
                                                            
                                                                                                    
                                                    else: 
                                                        #if any pdf changes    and z is number of change                                                
                                                        cunSheet1["A{}".format(x+1)].value = cCArray[x]
                                                        cunSheet1["C{}".format(x+1)].value = cCArray[x]
                                                        cunSheet1["D{}".format(x+1)].value = ": cCArray[{}]".format(x)
                                                        cunSheet1["E{}".format(x+1)].value = objDT
                                                        cunSheet1["F{}".format(x+1)].value = objMsg[x]
                                                        cunSheet1["G{}".format(x+1)].value = objFA[x]
                                                        cunSheet1["H{}".format(x+1)].value = "firstDurationRunAppTime = {},currentDurationRunAppTime = {}".format(firstDurationRunAppTime,currentDurationRunAppTime)
                                                        cunInv.save("appendReport.xlsx")
                                                        z = 1
                                                        while True:
                                                                if changesSheet["A{}".format(z)].value is None:
                                                                    #  if cunSheet1["A1"].value  is None:
                                                                        changesSheet["A{}".format(z)].value = "changes time {}".format(x)
                                                                        changesSheet["B{}".format(z)].value = "firstDurationRunAppTime = {},currentDurationRunAppTime = {}".format(firstDurationRunAppTime,currentDurationRunAppTime)

                                                                        changesSheet["C{}".format(z)].value = "cCArray[{}] = ".format(x)
                                                                        changesSheet["D{}".format(z)].value = cCArray[x]
                                                                        
                                                                    
                                                                        # print("dbFCArray[{}]= {}".format(x,dbFCArray[x]))
                                                                        
                                                                        changesSheet["E{}".format(z)].value = "dbFCArray[{}] = ".format(x)
                                                                        
                                                                        # changesSheet["F{}".format(z)].value = dbFCArray[x]
                                                                        changesSheet["G{}".format(z)].value = objDT  
                                                                        changesSheet["H{}".format(z)].value = objMsg[x]
                                                                        changesSheet["I{}".format(z)].value = objFA[x] 
                                                                        cunInv.save("appendReport.xlsx")
                                                                        cunInv.save("appendReport_changes.xlsx")
                                                                        break
                                                                else:
                                                                    z = z + 1
            
                                                        cFOutPut = "🌟 {}\n\n✍️ {}\n     {}\n\n {}\n\n {}\n\n{}".format(objMsg[x],fA,objFA[x],chAdd,objDT,cCArray[x])
                                                        print("fCArray[{}] = {}\ncCArray[{}] = {}\n{}\nfirstDurationRunAppTime = {}\ncurrentDurationRunAppTime = {}\nCONTAINER CHANGED!".format(x,dbFCArray[x],x,cCArray[x],objDT,firstDurationRunAppTime,currentDurationRunAppTime))
                                                        # fCArray[x] = cCArray[x]
                                                        dbFCArray[x] = cCArray[x]
                                                        tel.send_message(channel,cFOutPut)
            
                                                
                                        
                                elif startTime > nowPerSecond():
                                        secWait = startTime-nowPerSecond()
                                        print("now is {} middle night I'm gono sleep till morning for  {} ".format(datetime.datetime.now(),secConToTime(secWait)))
                                        time.sleep(secWait)

                                elif finishTime < nowPerSecond(): 
                                        secWaitTom = (wholeTime-nowPerSecond()) + startTime
                                        print("now is {} night I'm gono sleep till tomorrow for {} ".format(datetime.datetime.now(),secConToTime(secWaitTom)))
                                        time.sleep(secWaitTom)

                    # except:
                    #     print("we got issue in try section while ")
                    #     Email().sendingEmail()
                    #     f=open("appendReport.xlsx","rb")
                    #     bot.send_document(myId,f,caption = "we got company in @updateDeutschBotschaft2_bot")
                        
                    except:    
                        print("(heart)stay in loop ")


                # except:
                #        print("we got connection vpn problem")
                #        Email().sendingEmail()
                #        f=open("appendReport.xlsx","rb")
                #        bot.send_document(myId,f,caption = "we got company in @updateDeutschBotschaft2_bot")


                except:    
                       print("(main)stay in loop ")

def nowPerSecond():
                    hNow = datetime.datetime.now().hour * 60 * 60
                    mNow = datetime.datetime.now().minute * 60
                    sNow = datetime.datetime.now().second
                    return hNow + mNow + sNow

def secConToTime(seconds):
    h = round(seconds / 3600)
    m = round((seconds % 3600) / 60)
    s = (seconds % 3600) % 60
    return "{}:{}:{}".format(h,m,s)





class telegram_bot:
       
            def __init__(self,config):
    
                 self.token = self.read_token_from_config_file(config)
                #  print("token =",self.token)
                 self.base = "https://api.telegram.org/bot{}/".format(self.token)
                 
                 
            
            def send_message(self,chat_id,msg): 
                    url = self.base + "sendMessage?chat_id={}&text={}".format(chat_id, msg) 
                    # print("url = {}".format(url))
                    if msg is not None:
                        requests.get(url)
             
            def read_token_from_config_file(self,config):
                     parser = cfg.ConfigParser()
                     parser.read(config)
                     #print(parser.read(config))
                     return parser.get('creds', 'token')



            

class Email:
          


            def sendingEmail(self,email=None,password=None,send_to_email=None,subject=None,message=None,file_location=None):
                    email = 'ss4925727@gmail.com'
                    password = 'qsapyrqzvacpsjuj'
                    send_to_email = 'ss433230@gmail.com'
                    subject = 'we got company on @updateDeutscheBotschaft_bot'
                    message = 'we got company on @updateDeutscheBotschaft_bot in folder updateDeutschBotschaft mail file take look on attach'
                    file_location = 'appendReport.xlsx'
                    msg = MIMEMultipart()
                    msg['From'] = email
                    msg['To'] = send_to_email
                    msg['Subject'] = subject

                    msg.attach(MIMEText(message, 'plain'))

                    # Setup the attachment
                    filename = os.path.basename(file_location)
                    attachment = open(file_location, "rb")
                    part = MIMEBase('application', 'octet-stream')
                    part.set_payload(attachment.read())
                    encoders.encode_base64(part)
                    part.add_header('Content-Disposition', "attachment; filename= %s" % filename)

                    # Attach the attachment to the MIMEMultipart object
                    msg.attach(part)

                    server = smtplib.SMTP('smtp.gmail.com', 587)
                    server.starttls()
                    server.login(email, password)
                    text = msg.as_string()
                    server.sendmail(email, send_to_email, text)
                    server.quit()

if __name__ == '__main__':
    main()


#success
